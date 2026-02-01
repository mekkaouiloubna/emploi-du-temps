"""Fonctionnalités d'exportation pour les emplois du temps et les rapports"""
from reportlab.lib.pagesizes import letter, A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, PageBreak, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import time
import io

class TimetableExporter:
    """
    Exportateur d'emploi du temps vers différents formats (PDF, Excel, CSV).
    Gère la mise en forme et la structuration des données pour l'impression ou le téléchargement.
    """
    
    def __init__(self, timeslots, title="Emploi du temps universitaire"):
        """
        Initialise l'exportateur.
        
        Args:
            timeslots: Liste des créneaux horaires à exporter.
            title: Titre du document généré.
        """
        self.timeslots = timeslots
        self.title = title
    
    def export_to_pdf(self, filename=None):
        """
        Exporte l'emploi du temps au format PDF.
        Utilise ReportLab pour générer un document structuré avec tableaux.
        """
        if filename is None:
            filename = io.BytesIO()
        
        # Création du document PDF (format paysage pour mieux accueillir le tableau)
        doc = SimpleDocTemplate(filename, pagesize=landscape(A4))
        story = []
        
        # Ajout du titre
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#4361ee'),
            spaceAfter=30,
            alignment=1
        )
        story.append(Paragraph(self.title, title_style))
        story.append(Spacer(1, 0.2 * inch))
        
        # Création du tableau de l'emploi du temps
        table_data = self._create_table_data()
        table = Table(table_data, colWidths=[1.2*inch]*6)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4361ee')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
        ]))
        
        story.append(table)
        doc.build(story)
        
        return filename
    
    def export_to_excel(self, filename=None):
        """
        Exporte l'emploi du temps au format Excel.
        Utilise OpenPyXL pour créer un classeur avec mise en forme.
        """
        if filename is None:
            filename = io.BytesIO()
        
        # Création du classeur
        wb = Workbook()
        ws = wb.active
        ws.title = "Emploi du temps"
        
        # Ajout du titre
        ws['A1'] = self.title
        ws['A1'].font = Font(size=14, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="4361ee", end_color="4361ee", fill_type="solid")
        ws.merge_cells('A1:F1')
        
        # Ajout des en-têtes de colonnes
        headers = ['Heure', 'Lundi', 'Mardi', 'Mercredi', 'Jeudi', 'Vendredi']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="3a0ca3", end_color="3a0ca3", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Remplissage des données
        time_slots = self._get_unique_times()
        for row_idx, time_slot in enumerate(time_slots, 4):
            ws.cell(row=row_idx, column=1).value = str(time_slot)
            
            for day in range(5):
                slots_for_day = [s for s in self.timeslots 
                               if s.day_of_week == day and s.start_time == time_slot]
                if slots_for_day:
                    slot = slots_for_day[0]
                    cell_value = f"{slot.course.code}\n{slot.room.code}"
                    ws.cell(row=row_idx, column=day+2).value = cell_value
        
        wb.save(filename)
        return filename
    
    def export_to_csv(self, filename=None):
        """Exporte l'emploi du temps au format CSV simple"""
        import csv
        
        if filename is None:
            filename = io.StringIO()
        
        writer = csv.writer(filename)
        
        # Écriture de l'en-tête
        writer.writerow(['Cours', 'Salle', 'Jour', 'Heure début', 'Heure fin', 'Enseignants'])
        
        # Écriture des données
        for slot in self.timeslots:
            teachers = ', '.join([t.full_name for t in slot.course.teachers])
            writer.writerow([
                slot.course.code,
                slot.room.code,
                ['Lun', 'Mar', 'Mer', 'Jeu', 'Ven', 'Sam', 'Dim'][slot.day_of_week] if slot.day_of_week is not None else '',
                str(slot.start_time),
                str(slot.end_time),
                teachers
            ])
        
        return filename
    
    def _create_table_data(self):
        """Prépare les données pour le tableau PDF"""
        days = ['Lundi', 'Mardi', 'Mercredi', 'Jeudi', 'Vendredi']
        times = self._get_unique_times()
        
        # Ligne d'en-tête
        table_data = [['Heure'] + days]
        
        # Lignes de données
        for time_slot in times:
            row = [str(time_slot)]
            for day in range(5):
                slots = [s for s in self.timeslots 
                        if s.day_of_week == day and s.start_time == time_slot]
                if slots:
                    slot = slots[0]
                    row.append(f"{slot.course.code}\n{slot.room.code}")
                else:
                    row.append('')
            table_data.append(row)
        
        return table_data
    
    def _get_unique_times(self):
        """Récupère et trie les heures de début uniques des créneaux"""
        times = set()
        for slot in self.timeslots:
            times.add(slot.start_time)
        return sorted(list(times))
